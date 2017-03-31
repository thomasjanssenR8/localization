import openpyxl
from haversine import haversine
from itertools import combinations


def get_data():
    start_row = 4                                                                           # Get starting row
    end_row = 4
    stop = False                                                                            # Get ending row
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=1, max_col=1):
        if not stop:
            for cell in row:
                if not cell.value:
                    stop = True
                else:
                    end_row += 1
    end_row -= 1

    amount_of_bssids = end_row - start_row + 1
    print('Amount of BSSIDs: ' + str(amount_of_bssids))

    latitudes = []
    longitudes = []
    bssids = []

    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=1):   # put BSSIDs in list
        for cell in row:
            if cell.value:
                bssids.append(cell.value)
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=3, max_col=3):   # put latitudes in list
        for cell in row:
            if cell.value:
                latitudes.append(cell.value)
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=4, max_col=4):   # put longitudes in list
        for cell in row:
            if cell.value:
                longitudes.append(cell.value)

    read_data = [start_row, end_row, amount_of_bssids, bssids, latitudes, longitudes]
    return read_data


def write_combinations_to_file():
    row_index = end_row + 6

    i = 0  # Write combination number in column A
    for row in sheet.iter_rows(min_row=row_index, max_row=row_index+len(bssid_combs)-1, min_col=1, max_col=1):
        for cell in row:
            i += 1
            cell.value = i

    i = 0  # Write BSSID 1 in column B
    for row in sheet.iter_rows(min_row=row_index, max_row=row_index+len(bssid_combs)-1, min_col=2, max_col=2):
        for cell in row:
            cell.value = bssid_combs[i][0]
            i += 1

    i = 0  # Write BSSID 2 in column C
    for row in sheet.iter_rows(min_row=row_index, max_row=row_index+len(bssid_combs)-1, min_col=3, max_col=3):
        for cell in row:
            cell.value = bssid_combs[i][1]
            i += 1

    i = 0  # Write latitude of BSSID 1 in column D
    for row in sheet.iter_rows(min_row=row_index, max_row=row_index+len(bssid_combs)-1, min_col=4, max_col=4):
        for cell in row:
            cell.value = lat_combs[i][0]
            i += 1

    i = 0  # Write longitude of BSSID 1 in column E
    for row in sheet.iter_rows(min_row=row_index, max_row=row_index + len(bssid_combs) - 1, min_col=5, max_col=5):
        for cell in row:
            cell.value = long_combs[i][0]
            i += 1

    i = 0  # Write latitude of BSSID 2 in column F
    for row in sheet.iter_rows(min_row=row_index, max_row=row_index + len(bssid_combs) - 1, min_col=6, max_col=6):
        for cell in row:
            cell.value = lat_combs[i][1]
            i += 1

    i = 0  # Write longitude of BSSID 2 in column G
    for row in sheet.iter_rows(min_row=row_index, max_row=row_index + len(bssid_combs) - 1, min_col=7, max_col=7):
        for cell in row:
            cell.value = long_combs[i][1]
            i += 1


def calcMean():
    i = 0  # Write mean latitude of the 2 BSSIDs in column H
    row_index = end_row + 6
    for row in sheet.iter_rows(min_row=row_index, max_row=row_index + len(bssid_combs) - 1, min_col=8, max_col=8):
        for cell in row:
            mean_latitude = (lat_combs[i][0] + lat_combs[i][1]) / 2
            cell.value = mean_latitude
            mean_latitudes.append(mean_latitude)
            i += 1

    i = 0  # Write mean latitude of the 2 BSSIDs in column I
    row_index = end_row + 6
    for row in sheet.iter_rows(min_row=row_index, max_row=row_index + len(bssid_combs) - 1, min_col=9, max_col=9):
        for cell in row:
            mean_longitude = (long_combs[i][0] + long_combs[i][1]) / 2
            cell.value = mean_longitude
            mean_longitudes.append(mean_longitude)
            i += 1


def calc_error():
    gps_latitude = sheet['J1'].value
    gps_longitude = sheet['K1'].value
    gps_coordinate = (gps_latitude, gps_longitude)

    i = 0  # Write error in column J, using the haversine function
    row_index = end_row + 6
    for row in sheet.iter_rows(min_row=row_index, max_row=row_index + len(bssid_combs) - 1, min_col=10, max_col=10):
        for cell in row:
            mean_coordinate = (mean_latitudes[i], mean_longitudes[i])
            distance = haversine(gps_coordinate, mean_coordinate)
            cell.value = distance
            i += 1


file = 'bssids.xlsx'                                            # Load Excel sheet of a location (e.g. BAP1)
book = openpyxl.load_workbook(filename=file)
sheet = book.get_sheet_by_name('BAP' + input('Give the sheet number: BAP'))

[start_row, end_row, amount_of_bssids, bssids, latitudes, longitudes] = get_data()   # Load data from Excel sheet

bssid_combs = list(combinations(bssids, 2))                     # Get all possible combinations of 2 BSSIDs
lat_combs = list(combinations(latitudes, 2))                    # Get all the combinations of the measured latitudes
long_combs = list(combinations(longitudes, 2))                  # Get all the combinations of the measured longitudes

write_combinations_to_file()                                    # Write all possible combinations to Excel file

mean_latitudes = []
mean_longitudes = []
calcMean()                                                      # Calculate the mean coordinate of each pair of BSSIDs
calc_error()                                                     # Calculate distance between GPS and mean coordinate

book.save(file)                                                 # Save the data






